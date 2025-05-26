import os

import openpyxl
from asgiref.sync import sync_to_async
from django.conf import settings
from openpyxl.styles import Font, Alignment
from shop_app.models import Order


@sync_to_async
def export_order_to_excel(order: Order):
    try:
        excel_dir = os.path.join(settings.MEDIA_ROOT, 'orders_excel')
        os.makedirs(excel_dir, exist_ok=True)

        file_path = os.path.join(excel_dir, 'orders.xlsx')

        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Заказы"

            headers = [
                "Номер заказа", "Дата создания", "Время создания", "Дата оплаты", "Время оплаты",
                "Пользователь (ID)", "Имя пользователя", "Username",
                "Данные доставки", "Список товаров", "Общая сумма",
                "Статус оплаты", "ID платежа ЮKassa"
            ]
            sheet.append(headers)

            header_font = Font(bold=True)
            alignment = Alignment(horizontal='center', vertical='center')
            for col_num, _ in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = alignment

        order_items = list(order.items.select_related('product'))

        products_list = []
        for item in order_items:
            product_name = item.product.name if item.product else 'Удаленный товар'
            products_list.append(f"{product_name} ({item.quantity} шт. по {item.price_at_purchase}₽)")
        products_str = "\n".join(products_list)

        created_date_str = order.created_at.strftime("%Y-%m-%d") if order.created_at else ""
        created_time_str = order.created_at.strftime("%H:%M:%S") if order.created_at else ""
        paid_date_str = order.paid_at.strftime("%Y-%m-%d") if order.paid_at else ""
        paid_time_str = order.paid_at.strftime("%H:%M:%S") if order.paid_at else ""

        row_data = [
            order.id,
            created_date_str,
            created_time_str,
            paid_date_str,
            paid_time_str,
            order.user.id if order.user else 'N/A',
            order.user.first_name if order.user else 'N/A',
            order.user.username if order.user else 'N/A',
            order.delivery_info,
            products_str,
            float(order.total_amount),
            order.get_payment_status_display(),
            order.yookassa_payment_id if order.yookassa_payment_id else ''
        ]

        sheet.append(row_data)

        for row in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    continue
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(file_path)

    except Exception as e:
        print(f"[Excel Export Error] {e}")
